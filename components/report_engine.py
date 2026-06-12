from io import BytesIO
import copy
import json
import pathlib
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from components.systems_rating import calculate_systems_rating
from components.admin_value_resolver import resolve_admin_linked_value

BASE_DIR = pathlib.Path(__file__).resolve().parents[1]

# Client-approved alignment between NSP system names and Admin template names.
NSP_TO_ADMIN_SYSTEM = {
    "Digestive": "Digestive",
    "Intestinal": "Intestinal",
    "Immune/Lymphatic": "Immune",
    "Glandular/Endocrine": "Glandular",
    "Musculoskeletal": "Musculoskeletal",
}

ADMIN_TO_NSP_SYSTEM = {v: k for k, v in NSP_TO_ADMIN_SYSTEM.items()}

def admin_system_for_nsp(system_name):
    return NSP_TO_ADMIN_SYSTEM.get(system_name, system_name)

def nsp_system_for_admin(system_name):
    return ADMIN_TO_NSP_SYSTEM.get(system_name, system_name)


def _load_json(rel_path):
    return json.loads((BASE_DIR / rel_path).read_text(encoding="utf-8"))

def _score(value):
    if value in [None, "", "Select", "NA"]:
        return 0
    try:
        return int(value)
    except Exception:
        return 0

def _answer(value):
    return "" if value in [None, "Select"] else str(value)

def _question_maps():
    laf = _load_json("config/laf_questions.json")
    nsp1 = _load_json("config/nsp_page1_questions.json")
    nsp2 = _load_json("config/nsp_page2_questions.json")
    return (
        {q["code"]: q for q in laf},
        {q["code"]: q for q in nsp1},
        {q["code"]: q for q in nsp2},
    )



# --------------------------------------------------------------------
# v27: Report NSP source integrity helpers
# --------------------------------------------------------------------
def _count_answers(answers):
    return len([v for v in (answers or {}).values() if v not in [None, "", "Select"]])

def _latest_report_instance_id(db, member_id):
    """Find the best assessment instance for final report if UI did not set one."""
    instances = db.get("assessment_instances", {}).get(member_id, []) or []
    if not instances:
        return ""

    def rank(inst):
        status = str(inst.get("status", "")).lower()
        submitted = bool(inst.get("submitted_for_review"))
        finalized_like = status in ["finalized", "review_required", "submitted", "completed"]
        return (
            1 if finalized_like or submitted else 0,
            int(inst.get("instance_number", 0) or 0),
            str(inst.get("created_date", "")),
        )

    return sorted(instances, key=rank, reverse=True)[0].get("instance_id", "")

def prepare_report_db(db, member_id, selected_instance_id=None):
    """Return a report-safe db copy and diagnostics.

    Priority:
    1. selected_instance_id from UI/session
    2. latest submitted/finalized assessment instance
    3. legacy member-level nsp1_responses/nsp2_responses
    """
    safe_db = copy.deepcopy(db)
    selected_instance_id = selected_instance_id or _latest_report_instance_id(safe_db, member_id)

    legacy_nsp1 = safe_db.get("nsp1_responses", {}).get(member_id, {}) or {}
    legacy_nsp2 = safe_db.get("nsp2_responses", {}).get(member_id, {}) or {}

    source = "legacy_member_responses"
    inst_nsp1 = {}
    inst_nsp2 = {}

    if selected_instance_id:
        inst_resp = safe_db.get("assessment_instance_responses", {}).get(selected_instance_id, {}) or {}
        inst_nsp1 = inst_resp.get("nsp1", {}) or {}
        inst_nsp2 = inst_resp.get("nsp2", {}) or {}
        if _count_answers(inst_nsp1) or _count_answers(inst_nsp2):
            source = "assessment_instance_responses"
            safe_db.setdefault("nsp1_responses", {})[member_id] = inst_nsp1
            safe_db.setdefault("nsp2_responses", {})[member_id] = inst_nsp2

    admin_source = "legacy_member_admin_assessment"
    inst_admin = {}
    if selected_instance_id:
        inst_admin_record = safe_db.get("admin_assessments_by_instance", {}).get(selected_instance_id, {}) or {}
        if isinstance(inst_admin_record, dict):
            inst_admin = inst_admin_record.get("data", {}) or {}
        if isinstance(inst_admin, dict) and inst_admin:
            safe_db.setdefault("admin_assessments", {})[member_id] = inst_admin
            admin_source = "assessment_instance_admin_assessment"

    nsp1_final = safe_db.get("nsp1_responses", {}).get(member_id, {}) or {}
    nsp2_final = safe_db.get("nsp2_responses", {}).get(member_id, {}) or {}
    rows = calculate_systems_rating(nsp1_final, nsp2_final)
    digestive_score = next((r.get("Score", 0) for r in rows if r.get("System") == "Digestive"), 0)

    safe_db["_report_meta"] = {
        "member_id": member_id,
        "selected_instance_id": selected_instance_id or "",
        "nsp_source": source,
        "nsp1_answer_count": _count_answers(nsp1_final),
        "nsp2_answer_count": _count_answers(nsp2_final),
        "digestive_score": digestive_score,
        "legacy_nsp1_answer_count": _count_answers(legacy_nsp1),
        "legacy_nsp2_answer_count": _count_answers(legacy_nsp2),
        "instance_nsp1_answer_count": _count_answers(inst_nsp1),
        "instance_nsp2_answer_count": _count_answers(inst_nsp2),
        "admin_source": admin_source,
        "instance_admin_answer_count": _count_answers(inst_admin) if isinstance(inst_admin, dict) else 0,
    }
    return safe_db, safe_db["_report_meta"]

def report_data_diagnostics(db, member_id, selected_instance_id=None):
    safe_db, meta = prepare_report_db(db, member_id, selected_instance_id)
    rows = calculate_systems_rating(
        safe_db.get("nsp1_responses", {}).get(member_id, {}),
        safe_db.get("nsp2_responses", {}).get(member_id, {}),
    )
    return meta, rows

def _admin_rows(db, member_id):
    if not db.get("_report_meta"):
        db, _ = prepare_report_db(db, member_id)
    templates = _load_json("config/admin_templates.json")
    admin = db.get("admin_assessments", {}).get(member_id, {})
    nsp1 = db.get("nsp1_responses", {}).get(member_id, {})
    nsp2 = db.get("nsp2_responses", {}).get(member_id, {})
    laf = db.get("laf_responses", {}).get(member_id, {})

    rows = []
    for system, groups in templates.items():
        for group in [g for g in groups if not g.get("deleted")]:
            subheader = group.get("heading", "")
            for item in [x for x in group.get("items", []) if not x.get("deleted")]:
                label = item.get("label", "")
                linked_code = item.get("linked_code")
                key = f"{system}|{subheader}|{label}"
                stored = admin.get(system, {}).get(key, "Select")
                source = "Manual"
                answer = stored

                if linked_code:
                    answer, meta = resolve_admin_linked_value(item, nsp1=nsp1, nsp2=nsp2, laf=laf, stored=stored)
                    source = meta.get("source_label", f"Linked: {linked_code}")

                rows.append({
                    "System": system,
                    "NSP System": nsp_system_for_admin(system),
                    "Subheader": subheader,
                    "Question": label,
                    "Answer": _answer(answer),
                    "Score": _score(answer),
                    "Source": source,
                    "Linked Code": linked_code or "",
                })
    return rows

def _select_top_systems_from_nsp(nsp_system_rows):
    """Select top 3 systems by NSP score; include rank 4 only if tied with rank 3."""
    scored = [
        (row["System"], row["Score"])
        for row in nsp_system_rows
        if row.get("Score", 0) > 0
    ]
    scored.sort(key=lambda x: x[1], reverse=True)

    selected = scored[:3]
    if len(scored) >= 4 and len(selected) >= 3 and scored[3][1] == selected[-1][1]:
        selected = scored[:4]
    return selected, scored

def compute_summary(db, member_id):
    if not db.get("_report_meta"):
        db, _ = prepare_report_db(db, member_id)
    """Central report logic.

    NSP Systems Rating Score is the single source of truth for:
    - Partial Assessment system score
    - Final Assessment system score
    - Top 3 / Top 4 system selection

    Admin Assessment scores are used only to structure findings inside the selected systems.
    """
    nsp1_answers = db.get("nsp1_responses", {}).get(member_id, {})
    nsp2_answers = db.get("nsp2_responses", {}).get(member_id, {})
    nsp_system_rows = calculate_systems_rating(nsp1_answers, nsp2_answers)
    selected, sorted_systems = _select_top_systems_from_nsp(nsp_system_rows)
    selected_system_names = [system for system, _ in selected]
    selected_admin_system_names = [admin_system_for_nsp(system) for system in selected_system_names]
    nsp_score_lookup = {row["System"]: row["Score"] for row in nsp_system_rows}

    admin_rows = _admin_rows(db, member_id)

    all_significant = [
        r for r in admin_rows
        if r["Score"] in [2, 3]
    ]

    selected_significant = [
        r for r in all_significant
        if r["System"] in selected_admin_system_names or r.get("NSP System") in selected_system_names
    ]

    # Admin subheader totals are used only for secondary ordering/grouping.
    subheader_scores = {}
    for r in admin_rows:
        if not (r["System"] in selected_admin_system_names or r.get("NSP System") in selected_system_names):
            continue
        key = (r["System"], r["Subheader"])
        subheader_scores[key] = subheader_scores.get(key, 0) + r["Score"]

    system_rank = {system: idx for idx, (system, score) in enumerate(selected)}
    admin_system_rank = {admin_system_for_nsp(system): idx for idx, (system, score) in enumerate(selected)}
    selected_significant.sort(
        key=lambda r: (
            admin_system_rank.get(r["System"], system_rank.get(r.get("NSP System"), 999)),
            -r["Score"],  # 3 first, then 2
            -subheader_scores.get((r["System"], r["Subheader"]), 0),
            r["Subheader"],
            r["Question"],
        )
    )

    all_significant.sort(
        key=lambda r: (
            -nsp_score_lookup.get(r.get("NSP System", r["System"]), 0),
            r["System"],
            -r["Score"],
            r["Subheader"],
            r["Question"],
        )
    )

    subheader_summary = []
    for system, nsp_score in selected:
        subheaders = []
        for (sys_name, subheader), total in subheader_scores.items():
            if sys_name == system and total > 0:
                subheaders.append((subheader, total))
        subheaders.sort(key=lambda x: x[1], reverse=True)
        for subheader, total in subheaders:
            subheader_summary.append({
                "System": system,
                "NSP System Score": nsp_score,
                "Subheader": subheader,
                "Admin Subheader Score": total,
            })

    return {
        "nsp_system_rows": nsp_system_rows,
        "system_scores": sorted_systems,  # NSP system scores, not admin totals
        "selected_systems": selected,
        "selected_admin_systems": selected_admin_system_names,
        "subheader_summary": subheader_summary,
        "significant_findings": selected_significant,
        "all_2_3_findings": all_significant,
        "admin_rows": admin_rows,
        "nsp_score_lookup": nsp_score_lookup,
    }

def _new_workbook():
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"
    return wb

def _style_sheet(ws):
    header_fill = PatternFill("solid", fgColor="064E3B")
    header_font = Font(color="FFFFFF", bold=True)
    section_fill = PatternFill("solid", fgColor="FFF7E6")
    section_font = Font(color="064E3B", bold=True)
    thin = Side(style="thin", color="E9DFCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = border
            if cell.row == 1:
                cell.fill = header_fill
                cell.font = header_font
            if isinstance(cell.value, str) and cell.value.startswith("SECTION:"):
                cell.fill = section_fill
                cell.font = section_font

    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = max((len(str(c.value)) if c.value is not None else 0) for c in col)
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 14), 65)

def _append_table(wb, title, headers, rows):
    if "Report" in wb.sheetnames and wb["Report"].max_row == 1 and wb["Report"]["A1"].value is None:
        ws = wb["Report"]
        ws.title = title[:31]
    else:
        ws = wb.create_sheet(title[:31])
    ws.append(headers)
    for row in rows:
        ws.append(row)
    _style_sheet(ws)
    return ws

def _append_section(ws, title, headers=None, rows=None):
    ws.append([f"SECTION: {title}"])
    if headers:
        ws.append(headers)
    for row in rows or []:
        ws.append(row)
    ws.append([])

def _build_all_details_sheet(wb, db, member_id, summary):
    users = {u["id"]: u for u in db.get("users", [])}
    member = users.get(member_id, {})
    profile = db.get("profiles", {}).get(member_id, {})
    workflow = db.get("workflow", {}).get(member_id, {})

    laf_map, nsp1_map, nsp2_map = _question_maps()
    laf_answers = db.get("laf_responses", {}).get(member_id, {})
    nsp1_answers = db.get("nsp1_responses", {}).get(member_id, {})
    nsp2_answers = db.get("nsp2_responses", {}).get(member_id, {})
    body_mind_answers = db.get("body_mind_responses", {}).get(member_id, {})

    ws = wb.active
    ws.title = "All Details"

    report_meta = db.get("_report_meta", {})
    _append_section(ws, "Report Meta", ["Field", "Value"], [
        ["Generated On", datetime.now().strftime("%Y-%m-%d %H:%M")],
        ["Selected Instance", report_meta.get("selected_instance_id", "")],
        ["NSP Source", report_meta.get("nsp_source", "")],
        ["NSP1 Answers Used", report_meta.get("nsp1_answer_count", 0)],
        ["NSP2 Answers Used", report_meta.get("nsp2_answer_count", 0)],
        ["Digestive NSP Score", report_meta.get("digestive_score", 0)],
        ["Score Rule", "NSP Systems Rating Score is calculated from NSP Page 1 + NSP Page 2 and is used consistently in Partial and Final Reports."],
        ["Top System Rule", "Top 3 systems by NSP score. If rank 4 ties with rank 3, rank 4 is also included."],
        ["Admin Finding Rule", "Admin Assessment scores are used for findings; only 2 and 3 are shown in the final summary."],
    ])

    _append_section(ws, "Member Profile", ["Field", "Value"], [
        ["Name", member.get("name", "")],
        ["Email", member.get("email", "")],
        *[[k, v] for k, v in profile.items()],
    ])

    _append_section(ws, "Workflow", ["Field", "Value"], [[k, v] for k, v in workflow.items()])

    _append_section(ws, "NSP Systems Rating Table", ["No.", "System", "NSP Score"], [
        [r["No."], r["System"], r["Score"]] for r in summary["nsp_system_rows"]
    ])

    _append_section(ws, "LAF Responses", ["Section", "Question", "Answer"], [
        [laf_map.get(code, {}).get("section", ""), laf_map.get(code, {}).get("label", code), _answer(ans)]
        for code, ans in laf_answers.items()
    ])

    _append_section(ws, "NSP Page 1", ["No.", "Question", "Answer", "Score"], [
        [nsp1_map.get(code, {}).get("number", ""), nsp1_map.get(code, {}).get("text", code), _answer(ans), _score(ans)]
        for code, ans in nsp1_answers.items()
    ])

    _append_section(ws, "NSP Page 2", ["No.", "Question", "Answer", "Score"], [
        [nsp2_map.get(code, {}).get("number", ""), nsp2_map.get(code, {}).get("text", code), _answer(ans), _score(ans)]
        for code, ans in nsp2_answers.items()
    ])

    _append_section(ws, "Admin Assessment", ["System", "Subheader", "Question", "Answer", "Admin Score", "Source", "Linked Code"], [
        [r["System"], r["Subheader"], r["Question"], r["Answer"], r["Score"], r["Source"], r["Linked Code"]]
        for r in summary["admin_rows"]
    ])

    if body_mind_answers:
        try:
            body_mind_questions = _load_json("config/body_mind_questions.json")
            body_mind_map = {q["code"]: q for q in body_mind_questions}
            _append_section(ws, "Body Mind Connection", ["Section", "Question", "Answer"], [
                [body_mind_map.get(code, {}).get("section", ""), body_mind_map.get(code, {}).get("label", code), _answer(ans)]
                for code, ans in body_mind_answers.items()
            ])
        except Exception:
            pass

    _style_sheet(ws)

def build_full_admin_report(db, member_id):
    """Three-tab final report requested by client.

    Tab 1: All Details
    Tab 2: All 2 & 3 Elements
    Tab 3: Final Summary
    """
    wb = _new_workbook()
    summary = compute_summary(db, member_id)
    score_lookup = summary["nsp_score_lookup"]

    _build_all_details_sheet(wb, db, member_id, summary)

    _append_table(wb, "All 2 & 3 Elements", [
        "System", "NSP System Score", "Subheader", "Question", "Admin Score", "Answer", "Source", "Linked Code"
    ], [
        [
            r["System"],
            score_lookup.get(r.get("NSP System", r["System"]), 0),
            r["Subheader"],
            r["Question"],
            r["Score"],
            r["Answer"],
            r["Source"],
            r["Linked Code"],
        ]
        for r in summary["all_2_3_findings"]
    ])

    ws = wb.create_sheet("Final Summary")
    _append_section(ws, "Selected Top Systems", ["Rank", "System", "NSP System Score"], [
        [idx + 1, system, score] for idx, (system, score) in enumerate(summary["selected_systems"])
    ])
    _append_section(ws, "Final Structured Findings", [
        "System Rank", "System", "NSP System Score", "Subheader", "Question", "Admin Score", "Source", "Linked Code"
    ], [
        [
            next((idx + 1 for idx, (system, _) in enumerate(summary["selected_systems"]) if admin_system_for_nsp(system) == r["System"] or system == r.get("NSP System")), ""),
            r["System"],
            score_lookup.get(r.get("NSP System", r["System"]), 0),
            r["Subheader"],
            r["Question"],
            r["Score"],
            r["Source"],
            r["Linked Code"],
        ]
        for r in summary["significant_findings"]
    ])
    _style_sheet(ws)

    # Ensure exactly 3 tabs in preferred order
    for sheet_name in list(wb.sheetnames):
        if sheet_name not in ["All Details", "All 2 & 3 Elements", "Final Summary"]:
            del wb[sheet_name]
    wb._sheets = [wb["All Details"], wb["All 2 & 3 Elements"], wb["Final Summary"]]

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()

def build_summary_report(db, member_id):
    """Alias for the same three-tab final report, retained for existing page compatibility."""
    return build_full_admin_report(db, member_id)

def summary_preview_rows(db, member_id):
    summary = compute_summary(db, member_id)
    return summary["selected_systems"], summary["subheader_summary"], summary["significant_findings"]