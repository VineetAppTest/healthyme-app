from copy import copy
from functools import lru_cache
from io import BytesIO
import json
import pathlib

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from components.systems_rating import SYSTEM_ORDER, load_systems_rating_map

BASE_DIR = pathlib.Path(__file__).resolve().parents[1]


@lru_cache(maxsize=1)
def question_systems_map():
    """Return the common NSP question-code -> systems mapping in display order."""
    reverse_map = {}
    source_map = load_systems_rating_map()
    for system in SYSTEM_ORDER:
        for code in source_map.get(system, []):
            reverse_map.setdefault(code, []).append(system)
    return reverse_map


def systems_for_question(code):
    """Return a comma-separated system label for one NSP question code."""
    return ", ".join(question_systems_map().get(code, []))


@lru_cache(maxsize=2)
def _question_lookup(relative_path):
    questions = json.loads((BASE_DIR / relative_path).read_text(encoding="utf-8"))
    by_pair = {}
    by_text = {}
    for question in questions:
        code = question.get("code", "")
        number = str(question.get("number", "")).strip()
        text = str(question.get("text", "")).strip()
        if code and text:
            by_pair[(number, text)] = code
            by_text[text] = code
    return by_pair, by_text


def _copy_cell_style(source, target):
    if source.has_style:
        target.font = copy(source.font)
        target.fill = copy(source.fill)
        target.border = copy(source.border)
        target.alignment = copy(source.alignment)
        target.number_format = source.number_format
        target.protection = copy(source.protection)


def _add_systems_to_section(ws, section_title, question_file):
    section_row = next(
        (
            row_index
            for row_index in range(1, ws.max_row + 1)
            if ws.cell(row_index, 1).value == f"SECTION: {section_title}"
        ),
        None,
    )
    if section_row is None:
        return False

    header_row = section_row + 1
    headers = {
        str(ws.cell(header_row, column).value or "").strip(): column
        for column in range(1, ws.max_column + 1)
    }
    if "Systems" in headers:
        return True

    score_column = headers.get("Score")
    if score_column is None:
        return False

    systems_column = score_column + 1
    systems_header = ws.cell(header_row, systems_column, "Systems")
    _copy_cell_style(ws.cell(header_row, score_column), systems_header)

    by_pair, by_text = _question_lookup(question_file)
    row_index = header_row + 1
    while row_index <= ws.max_row:
        if all(
            ws.cell(row_index, column).value in [None, ""]
            for column in range(1, score_column + 1)
        ):
            break
        number = str(ws.cell(row_index, 1).value or "").strip()
        question = str(ws.cell(row_index, 2).value or "").strip()
        code = by_pair.get((number, question)) or by_text.get(question, "")
        systems_cell = ws.cell(row_index, systems_column, systems_for_question(code))
        _copy_cell_style(ws.cell(row_index, score_column), systems_cell)
        row_index += 1

    max_length = max(
        len(str(ws.cell(row, systems_column).value or ""))
        for row in range(header_row, row_index)
    )
    ws.column_dimensions[get_column_letter(systems_column)].width = min(max(max_length + 2, 20), 65)
    return True


def add_systems_column_to_final_report(report_bytes):
    """Add Systems beside Score in the two NSP sections of an existing final report."""
    workbook = load_workbook(BytesIO(report_bytes))
    if "All Details" not in workbook.sheetnames:
        return report_bytes

    worksheet = workbook["All Details"]
    page_1_updated = _add_systems_to_section(
        worksheet,
        "NSP Page 1",
        "config/nsp_page1_questions.json",
    )
    page_2_updated = _add_systems_to_section(
        worksheet,
        "NSP Page 2",
        "config/nsp_page2_questions.json",
    )
    if not (page_1_updated or page_2_updated):
        return report_bytes

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()
