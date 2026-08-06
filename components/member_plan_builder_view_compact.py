from __future__ import annotations

import html
import io
from typing import Any, Dict, List, Sequence

import pandas as pd
import streamlit as st

from components.current_member_plan import build_current_member_plan
from components.member_plan_presentation import (
    MEAL_TIMINGS,
    allocation_day_groups,
    meal_day_groups,
    profile_matches_or_filters,
    section_rows,
)
from components.pbm_core import clean
from components.recommendation_profile_viewer import (
    _render_view_profiles_css,
    load_profile_detail_readonly,
    load_profile_inventory,
)


@st.cache_data(ttl=60, show_spinner=False)
def _cached_inventory():
    return load_profile_inventory()


@st.cache_data(ttl=60, show_spinner=False)
def _cached_detail(profile_id: str):
    return load_profile_detail_readonly(profile_id)


@st.cache_data(ttl=30, show_spinner=False)
def _cached_current_plan(member_id: str):
    return build_current_member_plan(member_id)


def _profile_label(row: Dict[str, Any]) -> str:
    status = clean(row.get("status")).title() or "Unknown"
    start = clean(row.get("start_date")) or "No start date"
    return f"{clean(row.get('profile_name')) or 'Untitled'} · {status} · {start}"


def _member_label(row: Dict[str, Any]) -> str:
    return (
        clean(row.get("assigned_member_label"))
        or clean(row.get("assigned_member_id"))
        or "Unallocated"
    )


def _html_cell(value: object) -> str:
    text = html.unescape(str(value or ""))
    return "<br>".join(html.escape(line) for line in text.splitlines()) or "&mdash;"


def _meal_grid_cell(items: List[Dict[str, Any]], day_number: int, timing: str) -> str:
    groups = {
        clean(group.get("Timing")): group
        for group in meal_day_groups(items, day_number)
    }
    group = groups.get(timing, {})
    values = [
        clean(value)
        for value in (group.get("Meal"), group.get("Liquid"))
        if clean(value)
    ]
    return " + ".join(values)


def _render_meal_week_grid(items: List[Dict[str, Any]]) -> None:
    header_html = "".join(f"<th>{html.escape(slot)}</th>" for slot in MEAL_TIMINGS)
    rows: List[str] = []
    for day_number in range(1, 8):
        cells = "".join(
            f"<td>{_html_cell(_meal_grid_cell(items, day_number, slot))}</td>"
            for slot in MEAL_TIMINGS
        )
        rows.append(f"<tr><td>Day {day_number}</td>{cells}</tr>")
    st.markdown(
        """
<style id="mpb-weekly-plan-grid-v1">
.mpb-weekly-title{color:#064E3B;font-size:.92rem;font-weight:950;margin:.72rem 0 .28rem}
.mpb-plan-table-wrap{overflow:auto;border:1px solid #D8A84E;border-radius:12px;background:#fff;margin:.34rem 0 .78rem}
.mpb-plan-table{width:100%;border-collapse:collapse;font-size:.75rem;line-height:1.32}
.mpb-plan-table th{background:#FFF4DE;color:#064E3B;font-weight:900;text-align:center;padding:.45rem .42rem;border:1px solid #D8A84E;white-space:nowrap}
.mpb-plan-table td{color:#334155;font-weight:650;padding:.48rem .44rem;border:1px solid #E3C98E;vertical-align:top}
.mpb-plan-table td:first-child{color:#064E3B;font-weight:900;text-align:center;white-space:nowrap;background:#FFFDF8}
@media(max-width:760px){.mpb-plan-table{min-width:840px}}
</style>
""",
        unsafe_allow_html=True,
    )
    st.markdown(
        "<div class='mpb-weekly-title'>Meal</div>"
        "<div class='mpb-plan-table-wrap'><table class='mpb-plan-table'>"
        f"<thead><tr><th>Day</th>{header_html}</tr></thead>"
        f"<tbody>{''.join(rows)}</tbody></table></div>",
        unsafe_allow_html=True,
    )


def _model_rows(model: Dict[str, Any], domain: str) -> List[Dict[str, Any]]:
    partitions = dict((model or {}).get(domain) or {})
    rows = [dict(row or {}) for row in partitions.get("current") or []]
    rows.extend(dict(row or {}) for row in partitions.get("upcoming") or [])
    rows.sort(
        key=lambda row: (
            clean(row.get("start_date")),
            clean(row.get("exercise_name") or row.get("supplement_name") or row.get("title")).casefold(),
            clean(row.get("id")),
        )
    )
    return rows


def _snapshot(row: Dict[str, Any]) -> Dict[str, Any]:
    source_snapshot = row.get("source_snapshot")
    if isinstance(source_snapshot, dict):
        original = source_snapshot.get("source_original_snapshot")
        return dict(original) if isinstance(original, dict) else dict(source_snapshot)
    return {}


def _status_label(row: Dict[str, Any]) -> str:
    return (clean(row.get("effective_state") or row.get("status")) or "current").replace(
        "_", " "
    ).title()


def _date_range(row: Dict[str, Any]) -> str:
    start = clean(row.get("start_date")) or "No start"
    end = clean(row.get("end_date")) or "Open"
    return f"{start} to {end}"


def _render_flat_plan_table(
    title: str,
    rows: List[Dict[str, str]],
    headers: Sequence[str],
    empty_message: str,
) -> None:
    st.markdown(f"<div class='mpb-weekly-title'>{html.escape(title)}</div>", unsafe_allow_html=True)
    if not rows:
        st.info(empty_message)
        return
    header_html = "".join(f"<th>{html.escape(header)}</th>" for header in headers)
    body_html = "".join(
        "<tr>"
        + "".join(f"<td>{_html_cell(row.get(header))}</td>" for header in headers)
        + "</tr>"
        for row in rows
    )
    st.markdown(
        "<div class='mpb-plan-table-wrap'><table class='mpb-plan-table'>"
        f"<thead><tr>{header_html}</tr></thead><tbody>{body_html}</tbody></table></div>",
        unsafe_allow_html=True,
    )


def _exercise_plan_rows(model: Dict[str, Any]) -> List[Dict[str, str]]:
    rows: List[Dict[str, str]] = []
    for row in _model_rows(model, "exercise"):
        snapshot = _snapshot(row)
        rows.append(
            {
                "Exercise": clean(row.get("exercise_name") or row.get("title") or snapshot.get("title")) or "Exercise",
                "Reps/Duration": clean(row.get("duration_or_reps") or snapshot.get("duration_or_reps")) or "As advised",
                "Timing": clean(row.get("timing") or snapshot.get("timing")) or "As advised",
                "Dates": _date_range(row),
                "Status": _status_label(row),
                "Remarks": clean(row.get("instructions") or snapshot.get("instructions")),
            }
        )
    return rows


def _supplement_plan_rows(model: Dict[str, Any]) -> List[Dict[str, str]]:
    rows: List[Dict[str, str]] = []
    for row in _model_rows(model, "supplement"):
        snapshot = _snapshot(row)
        dosage = clean(row.get("dosage") or snapshot.get("dosage"))
        frequency = clean(row.get("frequency") or snapshot.get("frequency"))
        rows.append(
            {
                "Supplement": clean(
                    row.get("supplement_name")
                    or row.get("title")
                    or snapshot.get("supplement_name")
                    or snapshot.get("title")
                )
                or "Supplement",
                "Dosage": " · ".join(value for value in (dosage, frequency) if value) or "As advised",
                "Timing": clean(row.get("timing") or snapshot.get("timing")) or "As advised",
                "Dates": _date_range(row),
                "Status": _status_label(row),
                "Remarks": clean(row.get("instructions") or snapshot.get("instructions")),
            }
        )
    return rows


def _render_member_summary(profile: Dict[str, Any]) -> None:
    rows = [
        ("Member", _member_label(profile)),
        ("Meal Profile", clean(profile.get("profile_name")) or "Untitled"),
        ("Status", clean(profile.get("status")).title() or "Unknown"),
        ("Start Date", clean(profile.get("start_date")) or "Not set"),
        (
            "Health Concerns",
            ", ".join(clean(value) for value in profile.get("health_concerns") or [] if clean(value))
            or "Not set",
        ),
    ]
    body = "".join(
        f"<tr><th>{html.escape(label)}</th><td>{_html_cell(value)}</td></tr>"
        for label, value in rows
    )
    st.markdown(
        """
<style id="mpb-member-summary-table-v1">
.mpb-member-summary-wrap{border:1px solid #E3C98E;border-radius:14px;background:#FFFDF8;overflow:hidden;margin:.42rem 0 .78rem}
.mpb-member-summary-table{width:100%;border-collapse:collapse;font-size:.80rem;line-height:1.34}
.mpb-member-summary-table th{width:12rem;background:#FFF4DE;color:#064E3B;font-weight:950;text-align:left;padding:.48rem .56rem;border:1px solid #E3C98E}
.mpb-member-summary-table td{color:#334155;font-weight:720;padding:.48rem .56rem;border:1px solid #F0E3C5}
</style>
        """,
        unsafe_allow_html=True,
    )
    st.markdown(
        f"<div class='mpb-member-summary-wrap'><table class='mpb-member-summary-table'><tbody>{body}</tbody></table></div>",
        unsafe_allow_html=True,
    )


def _plan_sections(
    profile: Dict[str, Any],
    items: List[Dict[str, Any]],
    model: Dict[str, Any] | None,
) -> Dict[str, List[Dict[str, str]]]:
    start_date = clean(profile.get("start_date"))
    current_model = model or {}
    return {
        "Meals": section_rows(
            start_date=start_date,
            section_type="Meal",
            headers=("Timing", "Meal", "Liquid", "Remarks"),
            day_groups=lambda day: meal_day_groups(items, day),
        ),
        "Exercise": section_rows(
            start_date=start_date,
            section_type="Exercise",
            headers=("Timing", "Activity", "Reps/Duration", "Remarks"),
            day_groups=lambda day: allocation_day_groups(
                current_model, "exercise", start_date, day
            ),
        ),
        "Supplement": section_rows(
            start_date=start_date,
            section_type="Supplement",
            headers=("Timing", "Supplement", "Dosage", "Remarks"),
            day_groups=lambda day: allocation_day_groups(
                current_model, "supplement", start_date, day
            ),
        ),
    }


def _build_workbook(sections: Dict[str, List[Dict[str, str]]]) -> bytes:
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        for sheet_name, rows in sections.items():
            frame = pd.DataFrame(rows)
            if "Start Date" in frame:
                parsed_dates = pd.to_datetime(frame["Start Date"], errors="coerce")
                if parsed_dates.notna().all():
                    frame["Start Date"] = parsed_dates.dt.date
            frame.to_excel(writer, sheet_name=sheet_name, index=False)

        header_fill = PatternFill("solid", fgColor="FFF4DE")
        timing_fill = PatternFill("solid", fgColor="FFFCF5")
        header_border = Border(
            left=Side(style="thin", color="D8A84E"),
            right=Side(style="thin", color="D8A84E"),
            top=Side(style="thin", color="D8A84E"),
            bottom=Side(style="thin", color="D8A84E"),
        )
        body_border = Border(
            left=Side(style="thin", color="E3C98E"),
            right=Side(style="thin", color="E3C98E"),
            top=Side(style="thin", color="E3C98E"),
            bottom=Side(style="thin", color="E3C98E"),
        )
        preferred_widths = {
            "Start Date": 14,
            "Type": 14,
            "Day": 11,
            "Timing": 23,
            "Meal": 32,
            "Liquid": 25,
            "Activity": 30,
            "Duration/Sets": 20,
            "Reps/Duration": 20,
            "Supplement": 30,
            "Dosage": 22,
            "Remarks": 44,
        }
        for sheet in writer.book.worksheets:
            sheet.freeze_panes = "A2"
            sheet.auto_filter.ref = sheet.dimensions
            sheet.sheet_view.showGridLines = False
            sheet.page_setup.orientation = "landscape"
            sheet.page_setup.fitToWidth = 1
            sheet.page_setup.fitToHeight = 0
            sheet.sheet_properties.pageSetUpPr.fitToPage = True
            sheet.print_title_rows = "1:1"
            sheet.print_area = sheet.dimensions
            sheet.row_dimensions[1].height = 25

            for column_cells in sheet.columns:
                values = [str(cell.value or "") for cell in column_cells]
                header = str(column_cells[0].value or "")
                width = min(
                    52,
                    max(preferred_widths.get(header, 12), max(len(value) for value in values) + 2),
                )
                sheet.column_dimensions[column_cells[0].column_letter].width = width
                for row_index, cell in enumerate(column_cells, 1):
                    if row_index == 1:
                        cell.fill = header_fill
                        cell.font = Font(name="Arial", size=10, bold=True, color="064E3B")
                        cell.border = header_border
                        cell.alignment = Alignment(
                            horizontal="center", vertical="center", wrap_text=True
                        )
                    else:
                        cell.font = Font(name="Arial", size=10, color="334155")
                        cell.border = body_border
                        cell.alignment = Alignment(vertical="top", wrap_text=True)
                        if header == "Timing":
                            cell.fill = timing_fill
                            cell.font = Font(
                                name="Arial", size=10, bold=True, color="064E3B"
                            )
                        if header == "Start Date" and cell.value:
                            cell.number_format = "yyyy-mm-dd"

            for row_index in range(2, sheet.max_row + 1):
                line_count = max(
                    len(str(sheet.cell(row_index, column).value or "").splitlines())
                    for column in range(1, sheet.max_column + 1)
                )
                sheet.row_dimensions[row_index].height = max(20, 14 * line_count + 6)
    buffer.seek(0)
    return buffer.getvalue()


def _build_pdf(
    profile: Dict[str, Any],
    sections: Dict[str, List[Dict[str, str]]],
) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_LEFT
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import inch
    from reportlab.platypus import (
        LongTable,
        PageBreak,
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        TableStyle,
    )

    buffer = io.BytesIO()
    document = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=0.38 * inch,
        rightMargin=0.38 * inch,
        topMargin=0.38 * inch,
        bottomMargin=0.38 * inch,
        title=f"{clean(profile.get('profile_name')) or 'HealthyMe'} Member Plan",
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "HealthyMeTitle",
        parent=styles["Title"],
        fontName="Helvetica-Bold",
        fontSize=16,
        textColor=colors.HexColor("#064E3B"),
        alignment=TA_LEFT,
        spaceAfter=5,
    )
    section_style = ParagraphStyle(
        "HealthyMeSection",
        parent=styles["Heading2"],
        fontName="Helvetica-Bold",
        fontSize=11,
        textColor=colors.HexColor("#064E3B"),
        spaceBefore=7,
        spaceAfter=4,
    )
    cell_style = ParagraphStyle(
        "HealthyMeCell",
        parent=styles["BodyText"],
        fontName="Helvetica",
        fontSize=6.7,
        leading=8.3,
        textColor=colors.HexColor("#334155"),
    )

    def paragraph(value: object) -> Paragraph:
        escaped = html.escape(str(value or "—")).replace("\n", "<br/>")
        return Paragraph(escaped, cell_style)

    story = [
        Paragraph("HealthyMe Member Plan", title_style),
        Paragraph(
            " · ".join(
                value
                for value in (
                    clean(profile.get("assigned_member_label")),
                    clean(profile.get("profile_name")),
                    f"Start {clean(profile.get('start_date'))}"
                    if clean(profile.get("start_date"))
                    else "",
                )
                if value
            ),
            styles["BodyText"],
        ),
        Spacer(1, 5),
    ]

    for section_index, (section_name, rows) in enumerate(sections.items()):
        if section_index:
            story.append(PageBreak())
        story.append(Paragraph(section_name, section_style))
        headers = list(rows[0].keys()) if rows else []
        table_data = [[paragraph(header) for header in headers]]
        table_data.extend([[paragraph(row.get(header)) for header in headers] for row in rows])
        table = LongTable(
            table_data,
            repeatRows=1,
            colWidths=[0.84 * inch, 0.66 * inch, 0.52 * inch, 1.12 * inch, 2.05 * inch, 1.34 * inch, 2.15 * inch],
        )
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#FFF4DE")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#064E3B")),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("GRID", (0, 0), (-1, -1), 0.45, colors.HexColor("#D8A84E")),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 3),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 3),
                    ("TOPPADDING", (0, 0), (-1, -1), 3),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                ]
            )
        )
        story.extend([table, Spacer(1, 6)])

    def draw_page_number(canvas, _document) -> None:
        canvas.saveState()
        canvas.setFont("Helvetica", 7)
        canvas.setFillColor(colors.HexColor("#64748B"))
        canvas.drawRightString(
            landscape(A4)[0] - 0.38 * inch,
            0.18 * inch,
            f"Page {canvas.getPageNumber()}",
        )
        canvas.restoreState()

    document.build(
        story,
        onFirstPage=draw_page_number,
        onLaterPages=draw_page_number,
    )
    return buffer.getvalue()


def render_view_member_plan_compact() -> None:
    _render_view_profiles_css()
    st.markdown("<div class='hm-title'>View Member Plan</div>", unsafe_allow_html=True)

    ok, profiles, message = _cached_inventory()
    if not ok:
        st.error(message)
        return
    if not profiles:
        st.info("No Meal Profiles are available.")
        return

    profile_by_id = {clean(row.get("id")): row for row in profiles if clean(row.get("id"))}
    member_labels = {
        clean(row.get("assigned_member_id")): _member_label(row)
        for row in profiles
        if clean(row.get("assigned_member_id"))
    }
    health_concerns = sorted(
        {
            clean(concern)
            for row in profiles
            for concern in row.get("health_concerns") or []
            if clean(concern)
        }
    )

    profile_options = [""] + list(profile_by_id)
    member_options = [""] + list(member_labels)
    if st.session_state.get("mpb_view_profile_filter", "") not in profile_options:
        st.session_state["mpb_view_profile_filter"] = ""
    if st.session_state.get("mpb_view_member_filter", "") not in member_options:
        st.session_state["mpb_view_member_filter"] = ""

    controls = st.columns([0.40, 0.28, 0.32], gap="small")
    selected_profile_filter = controls[0].selectbox(
        "Meal Profile",
        profile_options,
        format_func=lambda value: (
            "All Meal Profiles" if not value else _profile_label(profile_by_id[value])
        ),
        key="mpb_view_profile_filter",
    )
    selected_member_filter = controls[1].selectbox(
        "Member",
        member_options,
        format_func=lambda value: "All Members" if not value else member_labels[value],
        key="mpb_view_member_filter",
    )
    selected_concerns = controls[2].multiselect(
        "Health Concerns",
        health_concerns,
        key="mpb_view_concern_filter",
        placeholder="All Health Concerns",
        help="Uses the Health Concern tags attached to each Meal Profile.",
    )

    matches = [
        row
        for row in profiles
        if profile_matches_or_filters(
            row,
            profile_id=selected_profile_filter,
            member_id=selected_member_filter,
            health_concerns=selected_concerns,
        )
    ]
    if not matches:
        st.info("No member plans match the selected OR filters.")
        return

    match_ids = [clean(row.get("id")) for row in matches if clean(row.get("id"))]
    result_key = "mpb_view_matching_plan"
    preferred = selected_profile_filter if selected_profile_filter in match_ids else match_ids[0]
    if st.session_state.get(result_key) not in match_ids:
        st.session_state[result_key] = preferred
    selected_id = (
        match_ids[0]
        if len(match_ids) == 1
        else st.selectbox(
            "Matching Meal Profile",
            match_ids,
            format_func=lambda value: (
                f"{_member_label(profile_by_id[value])} · {_profile_label(profile_by_id[value])}"
            ),
            key=result_key,
        )
    )

    detail_ok, profile, items, detail_message = _cached_detail(selected_id)
    if not detail_ok:
        st.error(detail_message)
        return

    member_id = clean(profile.get("assigned_member_id"))
    model: Dict[str, Any] | None = None
    if clean(profile.get("status")).lower() == "active":
        try:
            model = _cached_current_plan(member_id)
        except Exception as exc:
            st.error(f"Could not build the consolidated active plan: {exc}")
            return
        model_profile_id = clean((model.get("meal_profile") or {}).get("id"))
        if model_profile_id != selected_id:
            st.error(
                "Integrity check failed: the selected active Meal Profile does not match "
                "the member's consolidated current plan."
            )
            return

    plan_start = clean(profile.get("start_date"))
    _render_member_summary(profile)
    _render_meal_week_grid(items)
    _render_flat_plan_table(
        "Exercise",
        _exercise_plan_rows(model or {}),
        ("Exercise", "Reps/Duration", "Timing", "Dates", "Status", "Remarks"),
        "No Exercise allocations are visible for this member plan.",
    )
    _render_flat_plan_table(
        "Supplement",
        _supplement_plan_rows(model or {}),
        ("Supplement", "Dosage", "Timing", "Dates", "Status", "Remarks"),
        "No Supplement allocations are visible for this member plan.",
    )

    sections = _plan_sections(profile, items, model)
    file_stem = (
        f"{clean(profile.get('assigned_member_label')) or 'member'}_"
        f"{clean(profile.get('profile_name')) or 'plan'}"
    ).replace(" ", "_")
    download_columns = st.columns(2, gap="small")
    download_columns[0].download_button(
        "Download Excel",
        data=_build_workbook(sections),
        file_name=f"{file_stem}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
        key=f"mpb_view_download_xlsx_{selected_id}",
    )
    download_columns[1].download_button(
        "Download PDF",
        data=_build_pdf(profile, sections),
        file_name=f"{file_stem}.pdf",
        mime="application/pdf",
        use_container_width=True,
        key=f"mpb_view_download_pdf_{selected_id}",
    )