import copy
import hashlib
import json
import pathlib
from io import BytesIO

import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from components.db import (
    list_members,
    load_db,
    save_db_direct,
    update_member_response_with_audit,
)
from components.guards import require_admin
from components.ui_common import (
    apply_luxe_theme,
    card_end,
    card_start,
    inject_global_styles,
    render_back_to_top,
    render_page_nav,
    topbar,
    utility_logout_bar,
)


def v96_response_editor_success():
    st.success("Response updated. Please download a fresh final report.")


st.set_page_config(
    page_title="Response Editor",
    page_icon="💚",
    layout="wide",
    initial_sidebar_state="collapsed",
)
inject_global_styles()
apply_luxe_theme()
require_admin()
utility_logout_bar()

BASE = pathlib.Path(__file__).resolve().parents[1]
SUCCESS_KEY = "response_editor_success_message"
CLEANUP_KEY = "hm_response_editor_cleanup_keys"
VERSION_PREFIX = "hm_response_editor_version_"

FORM_STORES = {
    "LAF": ("laf_responses", "config/laf_questions.json"),
    "NSP Page 1": ("nsp1_responses", "config/nsp_page1_questions.json"),
    "NSP Page 2": ("nsp2_responses", "config/nsp_page2_questions.json"),
    "Body-Mind Page": ("body_mind_responses", "config/body_mind_questions.json"),
    "5 Admin Assessment Pages": ("admin_assessments", "config/admin_templates.json"),
}


def load_json(path):
    return json.loads((BASE / path).read_text(encoding="utf-8"))


def question_label(question):
    return question.get("label") or question.get("text") or question.get("code")


def _context_token(member_id, form_name, field_code):
    raw = f"{member_id}|{form_name}|{field_code}".encode("utf-8")
    return hashlib.sha1(raw).hexdigest()[:16]


def _version(token):
    key = f"{VERSION_PREFIX}{token}"
    return max(int(st.session_state.get(key, 1) or 1), 1)


def _bump_version(token):
    key = f"{VERSION_PREFIX}{token}"
    st.session_state[key] = _version(token) + 1


def _schedule_cleanup(*keys):
    st.session_state[CLEANUP_KEY] = tuple(str(key) for key in keys)


def _consume_cleanup():
    for key in st.session_state.pop(CLEANUP_KEY, ()) or ():
        st.session_state.pop(key, None)


def build_audit_excel(member, audit_rows):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Response Audit Log"
    headers = [
        "Timestamp",
        "Member",
        "Member Email",
        "Admin ID",
        "Form",
        "Field Code",
        "Old Value",
        "New Value",
        "Rationale",
    ]
    worksheet.append(headers)
    for item in audit_rows:
        worksheet.append(
            [
                item.get("timestamp", ""),
                member.get("name", ""),
                member.get("email", ""),
                item.get("admin_id", ""),
                item.get("form_name", ""),
                item.get("field_code", ""),
                item.get("old_value", ""),
                item.get("new_value", ""),
                item.get("rationale", ""),
            ]
        )
    style_excel(worksheet)
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def style_excel(worksheet):
    header_fill = PatternFill("solid", fgColor="064E3B")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="E9DFCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in worksheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = border
            if cell.row == 1:
                cell.fill = header_fill
                cell.font = header_font
    for column in worksheet.columns:
        column_letter = get_column_letter(column[0].column)
        max_length = max(
            (len(str(cell.value)) if cell.value is not None else 0)
            for cell in column
        )
        worksheet.column_dimensions[column_letter].width = min(
            max(max_length + 2, 14),
            60,
        )


def flatten_standard_questions(form_name, config_path):
    del form_name
    data = load_json(config_path)
    fields = []
    for question in data:
        status = "Inactive" if question.get("deleted") else "Active"
        fields.append(
            {
                "field_code": question.get("code", ""),
                "display": (
                    f"{status} — {question.get('code', '')} — "
                    f"{question_label(question)[:100]}"
                ),
                "question": question_label(question),
                "type": question.get("type", "text"),
                "options": question.get("options", []),
                "deleted": bool(question.get("deleted")),
            }
        )
    return fields


def flatten_admin_questions():
    templates = load_json("config/admin_templates.json")
    fields = []
    for system, groups in templates.items():
        for group in groups:
            heading = group.get("heading", "")
            for item in group.get("items", []):
                label = item.get("label", "")
                key = f"{system}|{heading}|{label}"
                status = (
                    "Inactive"
                    if item.get("deleted") or group.get("deleted")
                    else "Active"
                )
                fields.append(
                    {
                        "field_code": key,
                        "display": f"{status} — {system} > {heading} — {label[:100]}",
                        "question": label,
                        "system": system,
                        "subheader": heading,
                        "type": "select",
                        "options": ["NA", "1", "2", "3"],
                        "deleted": bool(
                            item.get("deleted") or group.get("deleted")
                        ),
                    }
                )
    return fields


def get_current_value(database, member_id, form_name, field):
    if form_name == "5 Admin Assessment Pages":
        system = field["system"]
        key = field["field_code"]
        return (
            database.setdefault("admin_assessments", {})
            .setdefault(member_id, {})
            .setdefault(system, {})
            .get(key, "")
        )
    store, _ = FORM_STORES[form_name]
    return database.setdefault(store, {}).setdefault(member_id, {}).get(
        field["field_code"],
        "",
    )


def set_current_value(database, member_id, form_name, field, new_value):
    if form_name == "5 Admin Assessment Pages":
        system = field["system"]
        key = field["field_code"]
        (
            database.setdefault("admin_assessments", {})
            .setdefault(member_id, {})
            .setdefault(system, {})
        )[key] = str(new_value)
    else:
        store, _ = FORM_STORES[form_name]
        database.setdefault(store, {}).setdefault(member_id, {})[
            field["field_code"]
        ] = str(new_value)


_consume_cleanup()
if st.session_state.pop(SUCCESS_KEY, None):
    v96_response_editor_success()

topbar(
    "Admin Response Editor",
    "Edit any member response, including blank/unanswered fields, and record rationale with timestamp.",
    "Audited admin correction",
)

members = list_members()
if not members:
    st.info("No members available.")
    st.stop()

member_map = {
    str(member.get("id", "")): member
    for member in members
    if member.get("id")
}
member_id = st.selectbox(
    "Select member",
    list(member_map.keys()),
    format_func=lambda value: (
        f"{value} — {member_map[value].get('name', '')} — "
        f"{member_map[value].get('email', '')}"
    ),
    key="hm_response_editor_member_id",
)
member = member_map.get(member_id, {})

database = load_db()
selected_form = st.selectbox(
    "Select response area",
    list(FORM_STORES.keys()),
    key="hm_response_editor_form",
)
_, config_path = FORM_STORES[selected_form]

if selected_form == "5 Admin Assessment Pages":
    fields = flatten_admin_questions()
else:
    fields = flatten_standard_questions(selected_form, config_path)

show_filter = st.radio(
    "Show fields",
    ["All fields", "Answered only", "Unanswered only"],
    horizontal=True,
    key=f"hm_response_editor_filter_{selected_form}",
)
filtered_fields = []
for field in fields:
    value = get_current_value(database, member_id, selected_form, field)
    answered = str(value).strip() not in ["", "Select", "None"]
    if show_filter == "Answered only" and not answered:
        continue
    if show_filter == "Unanswered only" and answered:
        continue
    filtered_fields.append(field)

card_start()
st.subheader("All responses / fields")
st.caption(
    f"Showing {len(filtered_fields)} of {len(fields)} fields for {selected_form}."
)
if not filtered_fields:
    st.info("No fields match the selected view.")
else:
    selected_field_index = st.selectbox(
        "Select field to edit",
        list(range(len(filtered_fields))),
        format_func=lambda index: filtered_fields[index]["display"],
        key=(
            "hm_response_editor_field_"
            f"{_context_token(member_id, selected_form, show_filter)}"
        ),
    )
    field = filtered_fields[selected_field_index]
    old_value = get_current_value(database, member_id, selected_form, field)
    token = _context_token(member_id, selected_form, field["field_code"])
    version = _version(token)
    value_key = f"hm_response_editor_value_{token}_v{version}"
    rationale_key = f"hm_response_editor_rationale_{token}_v{version}"

    st.markdown(f"**Question:** {field['question']}")
    st.markdown(f"**Current value:** `{old_value}`")

    if field["type"] in ["select", "scale"]:
        options = list(field.get("options", []))
        if field["type"] == "scale" and not options:
            options = [str(value) for value in range(1, 11)]
        options = ["Select"] + [value for value in options if value != "Select"]
        if selected_form.startswith("NSP") and "NA" not in options:
            options.insert(1, "NA")
        selected_index = options.index(old_value) if old_value in options else 0
        new_value = st.selectbox(
            "New value",
            options,
            index=selected_index,
            key=value_key,
        )
    elif field["type"] == "checkbox":
        new_value = st.checkbox(
            "New value",
            value=(str(old_value).lower() == "true" or old_value is True),
            key=value_key,
        )
    else:
        new_value = st.text_area(
            "New value",
            value=str(old_value),
            height=120,
            key=value_key,
        )

    rationale = st.text_area(
        "Rationale / note for change",
        placeholder="Mandatory. Example: Corrected after conversation with member.",
        height=100,
        key=rationale_key,
    )

    if st.button(
        "Save Edited Response with Audit Note",
        type="primary",
        use_container_width=True,
        key=f"hm_response_editor_save_{token}_v{version}",
    ):
        if str(new_value) == str(old_value):
            st.info("No value changed.")
        elif not rationale.strip():
            st.error("Rationale/note is mandatory for edited member responses.")
        else:
            try:
                candidate_database = copy.deepcopy(database)
                set_current_value(
                    candidate_database,
                    member_id,
                    selected_form,
                    field,
                    new_value,
                )
                save_db_direct(candidate_database)
                update_member_response_with_audit(
                    st.session_state.get("user_id", "admin"),
                    member_id,
                    selected_form,
                    field["field_code"],
                    old_value,
                    str(new_value),
                    rationale.strip(),
                )
            except Exception as exc:
                st.error(
                    "Unable to complete the audited response update. "
                    f"Your entered value and rationale have been retained. {exc}"
                )
            else:
                _schedule_cleanup(value_key, rationale_key)
                _bump_version(token)
                st.session_state[SUCCESS_KEY] = True
                st.rerun()
card_end()

card_start()
st.subheader("Audit log for this member")
audit = [
    item
    for item in load_db().get("response_audit_log", [])
    if item.get("member_id") == member_id
]
if not audit:
    st.info("No response edits recorded yet.")
else:
    st.download_button(
        "Download Member Response Audit Report",
        data=build_audit_excel(member, audit),
        file_name=(
            f"{member.get('name', 'member').replace(' ', '_')}_response_audit_log.xlsx"
        ),
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )
    for item in reversed(audit[-30:]):
        st.markdown(
            f"""
            **{item.get('timestamp', '')}** — {item.get('form_name', '')} / `{item.get('field_code', '')}`  
            Old: `{item.get('old_value', '')}` → New: `{item.get('new_value', '')}`  
            Rationale: {item.get('rationale', '')}
            """
        )
card_end()

render_page_nav(
    "Response Editor",
    back_page="pages/10_Admin_Dashboard.py",
    dashboard_page="pages/10_Admin_Dashboard.py",
    show_evaluation=False,
    show_dashboard=True,
    location="bottom",
)
render_back_to_top()
