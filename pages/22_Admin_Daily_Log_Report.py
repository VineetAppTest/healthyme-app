from datetime import date
from io import BytesIO

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

import components.ui_common as ui_common
from components.current_build import apply_current_build
from components.db import (
    get_daily_food_journal_day,
    get_daily_food_journal_days,
    get_daily_log_supervision_notes,
    get_meal_type_repository,
    list_members,
    queue_daily_log_reminder,
)
from components.guards import require_admin
from components.member_exercise_journal import list_member_exercise_logs, load_member_exercise_contract
from components.nutritionist_notes_h9a4 import create_structured_nutritionist_note, notes_for_journal_date
from components.ui_common import (
    apply_luxe_theme,
    card_end,
    card_start,
    compact_topbar,
    format_local_ts,
    inject_global_styles,
    render_back_to_top,
    render_page_nav,
    stat_grid,
    utility_logout_bar,
)

apply_current_build(ui_common)

st.set_page_config(page_title="Daily Food and Exercise Report", page_icon="💚", layout="wide", initial_sidebar_state="collapsed")
inject_global_styles()
apply_luxe_theme()
require_admin()
utility_logout_bar()

st.markdown(
    """
<style>
.block-container{padding-top:0!important;}
.utility-bar{margin-top:0!important;}
.hm-report-empty{background:#E9EEF5;border-radius:10px;padding:16px 18px;color:#334155;margin:.2rem 0 .9rem 0;}
.hm-report-divider{background:#111827;border-radius:8px;height:12px;margin:.75rem 0 .85rem 0;width:100%;}
.hm-guidance-heading{font-size:1.15rem;font-weight:900;color:#064E3B;margin:.35rem 0 .75rem 0;}
</style>
""",
    unsafe_allow_html=True,
)


def _safe_members():
    people = list_members()
    members = [m for m in people if str(m.get("role", "member") or "member").strip().lower() == "member"]
    if members:
        return members
    return [m for m in people if str(m.get("id", "")).lower() not in {"admin", "admin001"}]


def _meal_dict(meal):
    if isinstance(meal, dict):
        return meal
    if isinstance(meal, str):
        return {"food": meal}
    return {}


def _meal_label(key, meal):
    fallback = str(key or "meal").replace("_", " ").title()
    return str(_meal_dict(meal).get("label") or fallback)


def _meal_keys_for_day(day):
    keys = [(r["key"], r["label"]) for r in get_meal_type_repository()]
    known = {key for key, _label in keys}
    meals = day.get("meals", {}) if isinstance(day, dict) else {}
    if isinstance(meals, dict):
        for key, meal in meals.items():
            if key not in known:
                keys.append((key, _meal_label(key, meal)))
    return keys


def _structured_notes_text(member_id, log_date):
    values = []
    for row in notes_for_journal_date(member_id, log_date, include_archived=True):
        note = str(row.get("note_text", "") or "").strip()
        if note:
            values.append(f"{row.get('subject', 'Nutritionist Note')}: {note}")
    return " | ".join(values)


def _legacy_notes_text(member_id, log_date):
    values = []
    for row in get_daily_log_supervision_notes(member_id, limit=20, log_date=log_date):
        note = str(row.get("note", "") or "").strip()
        if note:
            values.append(note)
    return " | ".join(values)


def _other_fluids_summary(items):
    rows = []
    for idx, item in enumerate(items or [], start=1):
        if not isinstance(item, dict):
            continue
        rows.append(f"{idx}. {item.get('type') or 'Other Liquid'} | {item.get('time') or ''} | {item.get('quantity') or ''}")
    return "\n".join(rows) if rows else "—"


def _exercise_summary(member_id, log_date):
    rows = list_member_exercise_logs(member_id, log_date)
    if not rows:
        return "—"
    return "\n".join(
        f"{row.get('exercise_name','Exercise')} | {row.get('status','Not Started')} | {row.get('completion_time') or '-'} | {row.get('member_notes') or '-'}"
        for row in rows
    )


def _row_for_day(member_id, day):
    meals = day.get("meals", {}) if isinstance(day, dict) else {}
    meals = meals if isinstance(meals, dict) else {}
    log_date = day.get("date", "") if isinstance(day, dict) else ""
    breakfast = _meal_dict(meals.get("breakfast", {}))
    lunch = _meal_dict(meals.get("lunch", {}))
    dinner = _meal_dict(meals.get("dinner", {}))
    return {
        "Date": log_date,
        "Breakfast": breakfast.get("food", ""),
        "Lunch": lunch.get("food", ""),
        "Dinner": dinner.get("food", ""),
        "Water": day.get("water_litres", ""),
        "Other Fluids": _other_fluids_summary(day.get("other_fluids", [])),
        "Exercise": _exercise_summary(member_id, log_date),
        "Poop Rounds": day.get("poop_rounds", ""),
        "Poop Timings": ", ".join(str(x) for x in (day.get("poop_timings", []) or []) if str(x).strip()),
        "Feeling After Poop": day.get("feeling_after_poop", ""),
        "Notes": day.get("notes", ""),
        "Nutritionist Guidance": _structured_notes_text(member_id, log_date) or _legacy_notes_text(member_id, log_date),
    }


def _build_excel(member, days):
    wb = Workbook()
    ws = wb.active
    ws.title = "Daily Food and Exercise"
    ws.append(["Member", member.get("name", ""), "Email", member.get("email", "")])
    ws.append([])
    rows = [_row_for_day(member.get("id", ""), day) for day in days]
    headers = list(rows[0].keys()) if rows else list(_row_for_day(member.get("id", ""), {}).keys())
    ws.append(headers)
    for row in rows:
        ws.append([row.get(header, "") for header in headers])
    fill = PatternFill("solid", fgColor="064E3B")
    font = Font(color="FFFFFF", bold=True)
    border = Border(
        left=Side(style="thin", color="E9DFCC"), right=Side(style="thin", color="E9DFCC"),
        top=Side(style="thin", color="E9DFCC"), bottom=Side(style="thin", color="E9DFCC"),
    )
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = border
            if cell.row == 3:
                cell.fill = fill
                cell.font = font
    for col in range(1, ws.max_column + 1):
        ws.column_dimensions[ws.cell(row=3, column=col).column_letter].width = 24
    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


compact_topbar("Daily Food and Exercise Report", "", "Admin report")
members = _safe_members()
if not members:
    st.info("No members available.")
    st.stop()

options = [f"{m.get('id','')} — {m.get('name','Member')} — {m.get('email','')}" for m in members]
default_member = st.session_state.get("selected_daily_log_member_id") or st.session_state.get("selected_member_id")
default_index = next((idx for idx, option in enumerate(options) if default_member and option.startswith(f"{default_member} —")), 0)
selector_col, date_col = st.columns(2)
with selector_col:
    selected = st.selectbox("Select member", options, index=default_index)
member_id = selected.split(" — ")[0]
member = next(m for m in members if str(m.get("id", "")) == member_id)
days = get_daily_food_journal_days(member_id)
available_dates = [d.get("date") for d in days if d.get("date")]
default_date_str = available_dates[0] if available_dates else str(date.today())
try:
    default_date_obj = date.fromisoformat(default_date_str)
except Exception:
    default_date_obj = date.today()
with date_col:
    selected_date_obj = st.date_input("Select date for review", value=default_date_obj)
selected_date = selected_date_obj.isoformat()
selected_day = get_daily_food_journal_day(member_id, selected_date) or next((d for d in days if d.get("date") == selected_date), {"date": selected_date, "meals": {}})
exercise_contract = load_member_exercise_contract(member_id, member.get("email", ""))
exercise_logs = list_member_exercise_logs(member_id, selected_date)
structured_date_notes = notes_for_journal_date(member_id, selected_date, include_archived=True)
legacy_date_notes = get_daily_log_supervision_notes(member_id, limit=20, log_date=selected_date)

stat_grid([
    {"label": "Member", "value": member.get("name", ""), "note": "Selected member"},
    {"label": "Saved Days", "value": len(days), "note": "Food journal days"},
    {"label": "Assigned Exercises", "value": len(exercise_contract.get("exercises", [])), "note": "Current active profile / today"},
    {"label": "Exercise Logs", "value": len(exercise_logs), "note": "Selected review date"},
])

card_start()
st.subheader("All saved days")
if not days:
    st.markdown("<div class='hm-report-empty'>No daily food logs available.</div>", unsafe_allow_html=True)
else:
    rows = [_row_for_day(member_id, day) for day in days]
    st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)
    st.download_button(
        "Download Daily Food and Exercise Excel",
        data=_build_excel(member, days),
        file_name=f"{member.get('name','member').replace(' ','_')}_daily_food_exercise.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )
card_end()

card_start()
st.subheader(f"Food journal for {selected_date}")
selected_meals = selected_day.get("meals", {}) or {}
if not isinstance(selected_meals, dict) or not selected_meals:
    st.markdown("<div class='hm-report-empty'>No food journal available for this date.</div>", unsafe_allow_html=True)
else:
    meal_rows = []
    for key, label in _meal_keys_for_day(selected_day):
        meal = _meal_dict(selected_meals.get(key, {}))
        meal_rows.append({
            "Meal Type": label,
            "Time": meal.get("time", ""),
            "Food": meal.get("food", ""),
            "Portion Size": meal.get("portion_size", ""),
            "Mood": meal.get("mood", ""),
            "Energy": meal.get("energy", ""),
        })
    st.dataframe(pd.DataFrame(meal_rows), use_container_width=True, hide_index=True)
    st.markdown(f"**Water Intake:** {selected_day.get('water_litres','') or '-'}")
    st.markdown(f"**Other Fluids:** {_other_fluids_summary(selected_day.get('other_fluids', []))}")
    st.markdown(f"**Poop rounds:** {selected_day.get('poop_rounds','') or '-'}")
    st.markdown(f"**Feeling after poop:** {selected_day.get('feeling_after_poop','') or '-'}")
    st.markdown(f"**Overall Notes:** {selected_day.get('notes','') or '-'}")
    st.markdown("<div class='hm-report-divider'></div>", unsafe_allow_html=True)
card_end()

card_start()
st.subheader(f"Exercise for {selected_date}")
if exercise_logs:
    st.dataframe(pd.DataFrame([{
        "Exercise": row.get("exercise_name", ""),
        "Scheduled": row.get("scheduled_time", ""),
        "Difficulty": row.get("difficulty", ""),
        "Duration / Reps": row.get("duration_or_reps", ""),
        "Status": row.get("status", ""),
        "Completion Time": row.get("completion_time", ""),
        "Member Notes": row.get("member_notes", ""),
    } for row in exercise_logs]), use_container_width=True, hide_index=True)
elif exercise_contract.get("exercises"):
    st.caption("No member completion log exists for this date. Showing exercises from the active recommendation profile.")
    st.dataframe(pd.DataFrame([{
        "Exercise": row.get("name", ""),
        "Time of Day": row.get("timing_or_slot", ""),
        "Difficulty": row.get("difficulty", ""),
        "Duration / Reps": row.get("duration_or_reps", ""),
        "Equipment": row.get("equipment", ""),
        "Benefits": row.get("benefits", ""),
        "Instruction": row.get("instruction", ""),
    } for row in exercise_contract.get("exercises", [])]), use_container_width=True, hide_index=True)
else:
    st.markdown("<div class='hm-report-empty'>No exercise is available from the active recommendation profile.</div>", unsafe_allow_html=True)
card_end()

card_start()
st.subheader(f"Guidance linked to {selected_date}")
if structured_date_notes:
    for row in structured_date_notes:
        st.markdown(f"**{row.get('subject', 'Nutritionist Note')}**")
        st.write(row.get("note_text", ""))
        st.divider()
elif legacy_date_notes:
    for row in legacy_date_notes:
        st.markdown(f"**{format_local_ts(row.get('ts',''))}**")
        st.write(row.get("note", ""))
else:
    st.caption("No structured guidance is linked to this review date yet.")
card_end()

card_start()
st.subheader("Add Nutritionist Guidance")
note_type_label = st.radio("Guidance Type", ["Single Day", "Date Range", "General Guidance"], horizontal=True, key="h9a4_guidance_type_selector")
note_type = {"Single Day": "single_day", "Date Range": "date_range", "General Guidance": "general"}[note_type_label]
subject = "General Guidance" if note_type == "general" else "Nutritionist Note"
st.markdown(f"<div class='hm-guidance-heading'>{subject}</div>", unsafe_allow_html=True)
with st.form("h9a4_daily_report_structured_note", clear_on_submit=True):
    if note_type == "single_day":
        from_date = st.date_input("Note Date", value=selected_date_obj, key="h9a4_single_day_note_date")
        to_date = from_date
    elif note_type == "date_range":
        c1, c2 = st.columns(2)
        with c1:
            from_date = st.date_input("From Date", value=selected_date_obj, key="h9a4_range_from_date")
        with c2:
            to_date = st.date_input("To Date", value=selected_date_obj, key="h9a4_range_to_date")
    else:
        from_date = st.date_input("Guidance Date", value=selected_date_obj, key="h9a4_general_guidance_date")
        to_date = from_date
    note_text = st.text_area(subject, height=120)
    submitted = st.form_submit_button("Publish / Send Guidance", use_container_width=True)
if submitted:
    try:
        note = create_structured_nutritionist_note(
            member_id=member_id,
            member_name=member.get("name", ""),
            note_type=note_type,
            subject=subject,
            note_text=note_text,
            from_date=from_date,
            to_date=to_date,
            created_by=st.session_state.get("user_email") or st.session_state.get("oidc_email") or "admin",
        )
        st.success(f"Published {subject} {note.get('id')}.")
    except Exception as exc:
        st.error(str(exc))

if st.button("Send gentle Daily Log reminder", type="secondary", use_container_width=True):
    queue_daily_log_reminder(member_id)
    st.success("Reminder queued for the member and marked for email notification.")
card_end()

render_page_nav("Daily Food and Exercise Report", back_page="pages/10_Admin_Dashboard.py", dashboard_page="pages/10_Admin_Dashboard.py", show_evaluation=False, show_dashboard=True, location="bottom")
render_back_to_top()
