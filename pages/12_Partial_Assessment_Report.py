import streamlit as st
import copy
import json
import pathlib
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from components.guards import require_admin
from components.ui_common import inject_global_styles, apply_luxe_theme, topbar, card_start, card_end, utility_logout_bar, stat_grid
from components.db import load_db, get_form_response, get_workflow, get_profile
from components.systems_rating import calculate_systems_rating

st.set_page_config(page_title="Partial Assessment Report", page_icon="💚", layout="wide", initial_sidebar_state="collapsed")
inject_global_styles(); apply_luxe_theme(); require_admin(); utility_logout_bar()

mid = st.session_state.get("selected_member_id")
if not mid:
    st.switch_page("pages/11_Evaluation_Status.py")

BASE = pathlib.Path(__file__).resolve().parents[1]

def load_questions(rel):
    return json.loads((BASE / rel).read_text(encoding="utf-8"))

def answer_clean(v):
    if v in [None, "", "Select"]:
        return ""
    return str(v)

def score(v):
    if v in [None, "", "Select", "NA"]:
        return 0
    try:
        return int(v)
    except Exception:
        return 0

def make_excel(member, profile, workflow, laf_rows, nsp1_rows, nsp2_rows, systems_rating_rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "Member Profile"

    def style(ws):
        header_fill = PatternFill("solid", fgColor="064E3B")
        header_font = Font(color="FFFFFF", bold=True)
        thin = Side(style="thin", color="E9DFCC")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                cell.border = border
                if cell.row == 1:
                    cell.fill = header_fill
                    cell.font = header_font
        for col in ws.columns:
            col_letter = get_column_letter(col[0].column)
            max_len = max((len(str(c.value)) if c.value is not None else 0) for c in col)
            ws.column_dimensions[col_letter].width = min(max(max_len + 2, 14), 55)

    ws.append(["Field", "Value"])
    ws.append(["Member Name", member.get("name", "")])
    ws.append(["Email", member.get("email", "")])
    for k, v in profile.items():
        ws.append([k.replace("_", " ").title(), v])
    for k, v in workflow.items():
        ws.append([f"Workflow: {k}", v])
    style(ws)


    ws = wb.create_sheet("Systems Rating Table")
    ws.append(["SYSTEMS RATING TABLE", "", ""])
    ws.append(["No.", "System", "Score"])
    for row in systems_rating_rows:
        ws.append([row["No."], row["System"], row["Score"]])
    style(ws)

    for title, headers, rows in [
        ("LAF Responses", ["Section", "Question", "Answer"], laf_rows),
        ("NSP Page 1", ["No.", "Question", "Answer", "Score"], nsp1_rows),
        ("NSP Page 2", ["No.", "Question", "Answer", "Score"], nsp2_rows),
    ]:
        ws = wb.create_sheet(title)
        ws.append(headers)
        for r in rows:
            ws.append(r)
        style(ws)

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()

db = load_db()
selected_instance_id = st.session_state.get('selected_instance_id')
if selected_instance_id:
    db = copy.deepcopy(db)
    inst_resp = db.get('assessment_instance_responses', {}).get(selected_instance_id, {})
    if inst_resp:
        db.setdefault('nsp1_responses', {})[mid] = inst_resp.get('nsp1', db.get('nsp1_responses', {}).get(mid, {}))
        db.setdefault('nsp2_responses', {})[mid] = inst_resp.get('nsp2', db.get('nsp2_responses', {}).get(mid, {}))
users = {u["id"]: u for u in db.get("users", [])}
member = users.get(mid, {})
profile = get_profile(mid)
workflow = get_workflow(mid)

laf_questions = load_questions("config/laf_questions.json")
nsp1_questions = load_questions("config/nsp_page1_questions.json")
nsp2_questions = load_questions("config/nsp_page2_questions.json")

laf_map = {q["code"]: q for q in laf_questions}
nsp1_map = {q["code"]: q for q in nsp1_questions}
nsp2_map = {q["code"]: q for q in nsp2_questions}

laf = get_form_response("laf_responses", mid)
nsp1 = db.get("nsp1_responses", {}).get(mid, get_form_response("nsp1_responses", mid))
nsp2 = db.get("nsp2_responses", {}).get(mid, get_form_response("nsp2_responses", mid))
body_mind = db.get("body_mind_responses", {}).get(mid, {})
try:
    body_mind_questions = load_questions("config/body_mind_questions.json")
    body_mind_map = {q["code"]: q for q in body_mind_questions}
except Exception:
    body_mind_map = {}

laf_rows = []
for code, q in laf_map.items():
    ans = answer_clean(laf.get(code, ""))
    if ans:
        laf_rows.append([q.get("section", ""), q.get("label", code), ans])

nsp1_rows = []
for code, q in nsp1_map.items():
    ans = answer_clean(nsp1.get(code, ""))
    if ans:
        nsp1_rows.append([q.get("number", ""), q.get("text", code), ans, score(ans)])

nsp2_rows = []
for code, q in nsp2_map.items():
    ans = answer_clean(nsp2.get(code, ""))
    if ans:
        nsp2_rows.append([q.get("number", ""), q.get("text", code), ans, score(ans)])

systems_rating_rows = calculate_systems_rating(nsp1, nsp2)

topbar(
    "Partial Assessment Report",
    "View and download the member's LAF, NSP responses and Systems Rating Table before admin assessment.",
    "Admin partial report"
)

stat_grid([
    {"label": "LAF Fields", "value": len(laf_rows), "note": "Answered fields"},
    {"label": "NSP Page 1", "value": len(nsp1_rows), "note": "Answered items"},
    {"label": "NSP Page 2", "value": len(nsp2_rows), "note": "Answered items"},
    {"label": "Workflow", "value": workflow.get("workflow_status", "not_started").replace("_", " ").title(), "note": "Current status"},
])


card_start()
st.subheader("Systems Rating Table")
st.caption("Calculated from NSP Page 1 and NSP Page 2 responses.")
sr_header = st.columns([0.5, 3, 1])
with sr_header[0]:
    st.markdown("**No.**")
with sr_header[1]:
    st.markdown("**System**")
with sr_header[2]:
    st.markdown("**Score**")
for row in systems_rating_rows:
    c1, c2, c3 = st.columns([0.5, 3, 1])
    with c1:
        st.write(f"{row['No.']}.")
    with c2:
        st.write(row["System"])
    with c3:
        st.write(row["Score"])
card_end()

card_start()
st.subheader("Member Snapshot")
c1, c2, c3 = st.columns(3)
with c1:
    st.markdown(f"**Name:** {member.get('name', '')}")
    st.markdown(f"**Email:** {member.get('email', '')}")
with c2:
    st.markdown(f"**Age:** {profile.get('age', '')}")
    st.markdown(f"**Gender:** {profile.get('gender', '')}")
with c3:
    st.markdown(f"**Mobile:** {profile.get('mobile_number', profile.get('mobile_number', profile.get('phone', '')))}")
    st.markdown(f"**Country:** {profile.get('country', profile.get('country', profile.get('city', '')))}")
card_end()

card_start()
st.subheader("Viewable Report")
tabs = st.tabs(["LAF Summary", "NSP Page 1", "NSP Page 2", "Body-Mind"])

with tabs[0]:
    if not laf_rows:
        st.info("No LAF responses available yet.")
    else:
        # Group LAF by section and show in two columns for readability.
        sections = {}
        for section, question, ans in laf_rows:
            sections.setdefault(section or "General", []).append((question, ans))
        for section, rows in sections.items():
            st.markdown(f"### {section}")
            for i in range(0, len(rows), 2):
                cols = st.columns(2)
                for j, item in enumerate(rows[i:i+2]):
                    q, a = item
                    with cols[j]:
                        st.markdown(f"**{q}**")
                        st.write(a)

with tabs[1]:
    if not nsp1_rows:
        st.info("No NSP Page 1 responses available yet.")
    else:
        for i in range(0, len(nsp1_rows), 3):
            cols = st.columns(3)
            for j, row in enumerate(nsp1_rows[i:i+3]):
                no, q, a, s = row
                with cols[j]:
                    st.markdown(f"**{no}. {q}**")
                    st.write(f"Answer: {a}")
                    st.caption(f"Score: {s}")

with tabs[2]:
    if not nsp2_rows:
        st.info("No NSP Page 2 responses available yet.")
    else:
        for i in range(0, len(nsp2_rows), 3):
            cols = st.columns(3)
            for j, row in enumerate(nsp2_rows[i:i+3]):
                no, q, a, s = row
                with cols[j]:
                    st.markdown(f"**{no}. {q}**")
                    st.write(f"Answer: {a}")
                    st.caption(f"Score: {s}")
with tabs[3]:
    if not body_mind:
        st.info("Body-Mind Connection has not been filled or enabled yet.")
    else:
        for code, ans in body_mind.items():
            q = body_mind_map.get(code, {})
            st.markdown(f"**{q.get('label', code)}**")
            st.write(ans)

card_end()

card_start()
st.subheader("Download Partial Assessment Report")
safe_name = (member.get("name") or "member").replace(" ", "_").replace("/", "_")
xlsx = make_excel(member, profile, workflow, laf_rows, nsp1_rows, nsp2_rows, systems_rating_rows)
st.download_button(
    "Download Partial Assessment Report Excel",
    data=xlsx,
    file_name=f"{safe_name}_partial_assessment_report.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    use_container_width=True,
)
card_end()

if st.button("Back to Evaluation Status"):
    st.switch_page("pages/11_Evaluation_Status.py")