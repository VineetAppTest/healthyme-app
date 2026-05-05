import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from components.guards import require_admin
from components.ui_common import inject_global_styles, apply_luxe_theme, topbar, card_start, card_end, utility_logout_bar, stat_grid
from components.db import list_members, load_db
from components.assessment_instances import get_assessment_instances
from components.systems_rating import calculate_systems_rating

st.set_page_config(page_title="Comparative NSP Report", page_icon="💚", layout="wide", initial_sidebar_state="collapsed")
inject_global_styles(); apply_luxe_theme(); require_admin(); utility_logout_bar()

def score(v):
    if v in [None, "", "Select", "NA"]:
        return 0
    try:
        return int(v)
    except Exception:
        return 0

def style(ws):
    header_fill = PatternFill("solid", fgColor="064E3B")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="E9DFCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = border
            if cell.row == 1:
                cell.fill = header_fill
                cell.font = header_font
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        max_len = max((len(str(c.value)) if c.value is not None else 0) for c in col)
        ws.column_dimensions[letter].width = min(max(max_len + 2, 14), 55)

def build_excel(member, instances, system_df, q_df):
    wb = Workbook()
    ws = wb.active
    ws.title = "Systems Comparison"
    ws.append(list(system_df.columns))
    for row in system_df.itertuples(index=False):
        ws.append(list(row))
    style(ws)

    ws2 = wb.create_sheet("Question Comparison")
    ws2.append(list(q_df.columns))
    for row in q_df.itertuples(index=False):
        ws2.append(list(row))
    style(ws2)

    ws3 = wb.create_sheet("Instances")
    ws3.append(["Instance", "Type", "Requested Pages", "Status", "Submitted Date", "Due Date"])
    for inst in instances:
        ws3.append([
            inst.get("instance_number"),
            inst.get("instance_type"),
            ", ".join(inst.get("requested_pages", [])),
            inst.get("status"),
            inst.get("submitted_date"),
            inst.get("due_date"),
        ])
    style(ws3)

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()

topbar("Comparative NSP Report", "Compare NSP Systems Rating and question-level answers across assessment instances.", "Admin comparison")

members = list_members()
if not members:
    st.info("No members available.")
    st.stop()

selected = st.selectbox("Select member", [f"{m['id']} — {m['name']} — {m['email']}" for m in members])
member_id = selected.split(" — ")[0]
member = next(m for m in members if m["id"] == member_id)

db = load_db()
instances = get_assessment_instances(member_id)
responses = db.get("assessment_instance_responses", {})

completed_instances = [i for i in instances if i.get("submitted_for_review") or i.get("nsp1_completed") or i.get("nsp2_completed")]
completed_instances = sorted(completed_instances, key=lambda x: x.get("instance_number", 0))

stat_grid([
    {"label": "Member", "value": member["name"], "note": member["email"]},
    {"label": "Instances", "value": len(instances), "note": "All assessment rounds"},
    {"label": "Compared", "value": len(completed_instances), "note": "With available answers"},
    {"label": "Latest", "value": completed_instances[-1].get("instance_number") if completed_instances else "-", "note": "Latest instance"},
])

if not completed_instances:
    st.info("No NSP assessment instances available for comparison yet.")
    st.stop()

# Systems comparison
system_rows = {}
for inst in completed_instances:
    rid = inst["instance_id"]
    nsp1 = responses.get(rid, {}).get("nsp1", {})
    nsp2 = responses.get(rid, {}).get("nsp2", {})
    label = f"Instance {inst.get('instance_number')} - {inst.get('instance_type')}"
    for row in calculate_systems_rating(nsp1, nsp2):
        system_rows.setdefault(row["System"], {"System": row["System"]})
        system_rows[row["System"]][label] = row["Score"]

system_df = pd.DataFrame(list(system_rows.values())).fillna(0)

# Simple trend column if at least 2 instances
if len(completed_instances) >= 2:
    first_label = f"Instance {completed_instances[0].get('instance_number')} - {completed_instances[0].get('instance_type')}"
    latest_label = f"Instance {completed_instances[-1].get('instance_number')} - {completed_instances[-1].get('instance_type')}"
    def trend(row):
        try:
            if row[latest_label] < row[first_label]:
                return "Improved"
            if row[latest_label] > row[first_label]:
                return "Increased"
            return "No Change"
        except Exception:
            return "-"
    system_df["Trend vs Initial"] = system_df.apply(trend, axis=1)

# Question comparison
nsp1_questions = {q["code"]: q for q in __import__("json").loads((__import__("pathlib").Path(__file__).resolve().parents[1] / "config" / "nsp_page1_questions.json").read_text())}
nsp2_questions = {q["code"]: q for q in __import__("json").loads((__import__("pathlib").Path(__file__).resolve().parents[1] / "config" / "nsp_page2_questions.json").read_text())}
question_rows = {}
for inst in completed_instances:
    rid = inst["instance_id"]
    label = f"Instance {inst.get('instance_number')} - {inst.get('instance_type')}"
    for page, qmap in [("nsp1", nsp1_questions), ("nsp2", nsp2_questions)]:
        ans = responses.get(rid, {}).get(page, {})
        for code, q in qmap.items():
            key = f"{page}|{code}"
            question_rows.setdefault(key, {
                "Page": "NSP Page 1" if page == "nsp1" else "NSP Page 2",
                "No.": q.get("number", ""),
                "Question": q.get("text", code),
            })
            question_rows[key][label] = ans.get(code, "")

q_df = pd.DataFrame(list(question_rows.values())).fillna("")

card_start()
st.subheader("Systems Rating Comparison")
st.dataframe(system_df, use_container_width=True, hide_index=True)
card_end()

card_start()
st.subheader("Question-level Comparison")
st.dataframe(q_df, use_container_width=True, hide_index=True)
card_end()

card_start()
safe_name = member["name"].replace(" ", "_").replace("/", "_")
st.download_button(
    "Download Comparative NSP Report Excel",
    data=build_excel(member, completed_instances, system_df, q_df),
    file_name=f"{safe_name}_comparative_nsp_report.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    use_container_width=True,
)
card_end()

if st.button("Back to Dashboard"):
    st.switch_page("pages/10_Admin_Dashboard.py")