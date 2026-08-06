from __future__ import annotations

import io
from pathlib import Path

from openpyxl import load_workbook

from components.member_plan_builder_view_compact import (
    _build_pdf,
    _build_workbook,
    _plan_sections,
)
from components.member_plan_builder_allocation_common import allocation_choice_map
from components.member_plan_presentation import (
    allocation_day_groups,
    meal_day_groups,
    profile_matches_or_filters,
    split_timings,
)
from components.pbm_core import MEAL_SLOTS
from components.recommendation_profile_viewer import _meal_cells


ROOT = Path(__file__).resolve().parents[1]


def test_meals_are_clubbed_by_timing_in_chronological_order() -> None:
    items = [
        {"item_type": "meal", "day_number": 1, "slot_name": "Lunch", "item_order": 2, "reference_label": "Dal"},
        {"item_type": "meal", "day_number": 1, "slot_name": "Breakfast", "item_order": 1, "reference_label": "Moong Chilla"},
        {"item_type": "meal", "day_number": 1, "slot_name": "Lunch", "item_order": 1, "reference_label": "Paneer Salad", "portion": "1 bowl"},
    ]
    groups = meal_day_groups(items, 1)
    assert [group["Timing"] for group in groups] == ["Breakfast", "Lunch"]
    assert groups[1]["Meal"] == "Paneer Salad - 1 bowl + Dal"


def test_meal_slots_remove_wake_up_early_morning_and_use_food_portion_cells() -> None:
    assert "Wake-up / Early Morning" not in MEAL_SLOTS
    assert "Early Morning" not in MEAL_SLOTS
    items = [
        {"item_type": "meal", "day_number": 1, "slot_name": "Breakfast", "item_order": 1, "reference_label": "Oats", "portion": "1 bowl"},
        {"item_type": "meal", "day_number": 1, "slot_name": "Breakfast", "item_order": 2, "reference_label": "Apple", "portion": "1"},
    ]
    timing, meal, liquid, remarks = _meal_cells(items, 1)
    assert timing == "Breakfast"
    assert meal == "Oats - 1 bowl + Apple - 1"
    assert liquid == ""
    assert remarks == ""


def test_exercise_edit_choice_uses_reps_duration_start_end_cell_format() -> None:
    choices = allocation_choice_map(
        [
            {
                "exercise_name": "Cat-Cow Stretch",
                "start_date": "2026-08-06",
                "end_date": "2026-08-13",
                "status": "active",
                "source_snapshot": {"duration_or_reps": "10 reps"},
            }
        ],
        name_fields=("exercise_name", "title"),
        detail_fields=("duration_or_reps",),
        include_status=False,
        separator=" | ",
        date_format="%d %b %Y",
    )
    assert list(choices) == ["Cat-Cow Stretch | 10 reps | 06 Aug 2026 | 13 Aug 2026"]


def test_exercise_and_supplements_are_clubbed_and_sorted_by_timing() -> None:
    model = {
        "exercise": {
            "current": [
                {"exercise_name": "Squats", "timing": "Morning", "source_snapshot": {"duration_or_reps": "20-30"}},
                {"exercise_name": "Brisk Walk", "timing": "Morning", "source_snapshot": {"duration_or_reps": "30 min"}},
                {"exercise_name": "Stretch", "timing": "Evening", "source_snapshot": {"duration_or_reps": "10 min"}},
            ]
        },
        "supplement": {
            "current": [
                {"supplement_name": "Magnesium", "timing": "Evening, Morning", "dosage": "400", "frequency": "Once"},
                {"supplement_name": "Potassium", "timing": "Morning", "dosage": "100", "frequency": "Twice"},
            ]
        },
    }
    exercise = allocation_day_groups(model, "exercise", "2026-08-05", 1)
    assert [group["Timing"] for group in exercise] == ["Morning", "Evening"]
    assert exercise[0]["Activity"] == "Squats\nBrisk Walk"
    assert exercise[0]["Reps/Duration"] == "20-30\n30 min"
    supplements = allocation_day_groups(model, "supplement", "2026-08-05", 1)
    assert [group["Timing"] for group in supplements] == ["Morning", "Evening"]
    assert supplements[0]["Supplement"] == "Magnesium\nPotassium"
    assert split_timings("Evening, Morning, None") == ["Morning", "Evening", "None"]


def test_profile_member_and_concern_filters_use_or_logic() -> None:
    profile = {
        "id": "profile-1",
        "assigned_member_id": "member-1",
        "health_concerns": ["Gut Health"],
    }
    assert profile_matches_or_filters(profile, profile_id="other", member_id="member-1")
    assert profile_matches_or_filters(profile, profile_id="profile-1", health_concerns=["Sleep"])
    assert profile_matches_or_filters(profile, health_concerns=["Gut Health"])
    assert not profile_matches_or_filters(profile, member_id="member-2", health_concerns=["Sleep"])


def test_excel_and_pdf_exports_contain_all_three_screen_sections() -> None:
    profile = {
        "profile_name": "Balanced Plan",
        "assigned_member_label": "Test Member",
        "start_date": "2026-08-05",
    }
    items = [
        {"item_type": "meal", "day_number": 1, "slot_name": "Lunch", "reference_label": "Moong Chilla"}
    ]
    sections = _plan_sections(profile, items, {})
    assert list(sections) == ["Meals", "Exercise", "Supplement"]
    assert "Reps/Duration" in sections["Exercise"][0]
    workbook = _build_workbook(sections)
    pdf = _build_pdf(profile, sections)
    assert workbook.startswith(b"PK")
    assert pdf.startswith(b"%PDF")

    loaded = load_workbook(io.BytesIO(workbook))
    assert loaded.sheetnames == ["Meals", "Exercise", "Supplement"]
    for sheet in loaded.worksheets:
        assert sheet.freeze_panes == "A2"
        assert sheet.auto_filter.ref == sheet.dimensions
        assert sheet.sheet_view.showGridLines is False
        assert sheet["A2"].number_format == "yyyy-mm-dd"
        assert sheet["A1"].fill.fgColor.rgb.endswith("FFF4DE")
        assert sheet["A1"].font.bold is True


def test_setup_and_allocation_boundaries_follow_repository_model() -> None:
    setup = (ROOT / "components/member_plan_builder_setup.py").read_text(encoding="utf-8")
    meals = (ROOT / "components/member_plan_builder_meals_compact.py").read_text(encoding="utf-8")
    allocations = (ROOT / "components/member_plan_builder_allocation_common.py").read_text(encoding="utf-8")
    body_mind = (ROOT / "pages/19_Body_Mind_Connection.py").read_text(encoding="utf-8")
    assert '"Clone Meal Profile"' in setup
    assert '"assigned_member_id": ""' in setup
    assert '"Meal Profile"' in meals and '"Member"' in meals and '"Publish"' in meals
    assert "load_active_profiles()" in allocations
    assert '"Member"' in allocations
    assert '"Member Plan"' not in allocations
    assert "st.columns([1, 2, 1])" in body_mind
